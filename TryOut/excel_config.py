from openpyxl.styles import (Border, Side, Color, PatternFill, Font, Alignment, NamedStyle)


card_fonts = {
    'normal':           Font(name='Calibri', color=Color(rgb='000000'), size=11, bold=False, italic=False),
    'normal_bold':      Font(name='Calibri', color=Color(rgb='000000'), size=11, bold=True, italic=False),
    'normal_italic':    Font(name='Calibri', color=Color(rgb='000000'), size=11, bold=False, italic=True),
    'bold_blue':        Font(name='Calibri', color=Color(rgb='366092'), size=11, bold=True, italic=False),
    'bold_italic_blue': Font(name='Calibri', color=Color(rgb='366092'), size=11, bold=True, italic=True),
   }

card_borders = {
    "basic_left": Border(left=Side(border_style="thick", color=Color(rgb='A6A6A6')), right=None, top=None, bottom=None),
    "basic_right": Border(left=None, right=Side(border_style="thick", color=Color(rgb='A6A6A6')), top=None, bottom=None,),
    "basic_top": Border(left=None, right=None, top=Side(border_style="thick", color=Color(rgb='A6A6A6')), bottom=None,),
    "basic_bottom": Border(left=None, right=None, top=None, bottom=Side(border_style="thick", color=Color(rgb='A6A6A6'))),
   }

#
# fills = {
#     "title_basic": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00FAFAF4")),
#
#     "chapter_line": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00E4E2EC")),
#     "collection_line": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00ECDFEB")),
#     "section_line": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00E7ECDF")),
#     "subsection_line": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00DFECEB")),
#     "table_line": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00EEF3F8")),
#     "quote_line": PatternFill(patternType=None, fill_type=None, fgColor=Color(rgb="00FFFFFF")),
#
#     "extension_quotes": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00F9FAFA")),
#     "title_attributes": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00F1F1F9")),
#     "title_parameters": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00EFF6F2")),
#
# }
#
#
# alignments = {
#     "title_basic": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
#
#     "chapter_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
#     "collection_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
#     "section_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
#     "subsection_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
#     "table_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
#     "quote_line": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
#
#     "extension_quotes": Alignment(horizontal='left', vertical='bottom', wrap_text=True, shrink_to_fit=False, indent=0),
#     "title_attributes": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
#     "title_parameters": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
#
#     "right_alignment": Alignment(horizontal='right', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
#     "center_alignment": Alignment(horizontal='center', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
#
# }
#
# styles = {
#     "title_basic": NamedStyle(name="title_basic", font=fonts["title_basic"], border=borders["title_basic"],
#                               fill=fills["title_basic"], alignment=alignments["title_basic"]),
#
#     'chapter_line': NamedStyle(name="chapter_line", font=fonts["chapter_line"], border=borders["chapter_line"],
#                                fill=fills["chapter_line"], alignment=alignments["chapter_line"]),
#
#     'collection_line': NamedStyle(name="collection_line", font=fonts["collection_line"],
#                                   border=borders["collection_line"],
#                                   fill=fills["collection_line"], alignment=alignments["collection_line"]),
#
#     'section_line': NamedStyle(name="section_line", font=fonts["section_line"], border=borders["section_line"],
#                                fill=fills["section_line"], alignment=alignments["section_line"]),
#
#     'subsection_line': NamedStyle(name="subsection_line", font=fonts["subsection_line"],
#                                   border=borders["subsection_line"],
#                                   fill=fills["subsection_line"], alignment=alignments["subsection_line"]),
#
#     'table_line': NamedStyle(name="table_line", font=fonts["table_line"], border=borders["table_line"],
#                                   fill=fills["table_line"], alignment=alignments["table_line"]),
#
#     'quote_line': NamedStyle(name="quote_line", font=fonts["quote_line"], border=borders["quote_line"],
#                              fill=fills["quote_line"], alignment=alignments["quote_line"]),
#
#
#     'extension_quotes': NamedStyle(name="extension_quotes", font=fonts["extension_quotes"],
#                                    border=borders["extension_quotes"],
#                                    fill=fills["extension_quotes"], alignment=alignments["extension_quotes"]),
#
#     'title_attributes': NamedStyle(name="title_attributes", font=fonts["title_attributes"],
#                                    border=borders["title_attributes"],
#                                    fill=fills["title_attributes"], alignment=alignments["title_attributes"]),
#
#
#
#     'title_parameters': NamedStyle(name="title_parameters", font=fonts["title_parameters"],
#                                    border=borders["title_parameters"],
#                                    fill=fills["title_parameters"], alignment=alignments["title_parameters"]),
#
# }