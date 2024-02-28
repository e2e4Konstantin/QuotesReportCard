from openpyxl.worksheet import worksheet
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import DEFAULT_FONT
from openpyxl.styles import NamedStyle
import openpyxl
from icecream import ic

from excel_tools.config_data import Quote, Attribute, Option
from files_features import file_unused, output_message_exit
from excel_tools.excel_config import card_fonts, card_borders, card_alignments, card_styles


def _basic_header_output(sheet: worksheet):
    sheet.append(["Quote cards."])
    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 60
    sheet.column_dimensions['D'].width = 40


def _styles_init(workbook: openpyxl.Workbook, styles: dict[str:NamedStyle]):
    """ Добавляет в список стилей книги переданные в словаре стили. """
    for style in styles:
        if styles[style].name not in workbook.style_names:
            workbook.add_named_style(styles[style])
    ic(workbook.style_names)


def _quote_card_output(sheet: worksheet, quote_data: Quote, row: int) -> int:
    sheet.cell(row=row, column=column_index_from_string('B')).value = quote_data.code
    sheet.cell(row=row, column=column_index_from_string('C')).value = quote_data.title
    sheet.cell(row=row, column=column_index_from_string('D')).value = quote_data.measure

    sheet.row_dimensions[row].height = 30
    sheet.cell(row, column_index_from_string('B')).style = card_styles['h1']
    sheet.cell(row, column_index_from_string('B')).font = card_fonts['normal_bold']
    sheet.cell(row, column_index_from_string('B')).border = card_borders['quote_code']

    sheet.cell(row, column_index_from_string('C')).style = card_styles['h1']
    sheet.cell(row, column_index_from_string('C')).alignment = card_alignments['left_wrap']

    sheet.cell(row, column_index_from_string('D')).style = card_styles['h1']
    sheet.cell(row, column_index_from_string('D')).border = card_borders['h1_c']

    row += 1
    sheet.cell(row=row, column=column_index_from_string('B')).value = quote_data.table
    sheet.cell(row=row, column=column_index_from_string('C')).value = 'Атрибуты'
    sheet.cell(row=row, column=column_index_from_string('D')).value = ''

    sheet.cell(row, column_index_from_string('B')).style = card_styles['txt_gray']
    sheet.cell(row=row, column=column_index_from_string('C')).style = card_styles['h2']
    sheet.cell(row=row, column=column_index_from_string('D')).style = card_styles['h2']
    sheet.cell(row, column_index_from_string('D')).border = card_borders['txt_c']

    row += 1
    for attribute in quote_data.attributes:
        sheet.cell(row=row, column=column_index_from_string('B')).value = ''
        sheet.cell(row=row, column=column_index_from_string('C')).value = attribute.name
        sheet.cell(row=row, column=column_index_from_string('D')).value = attribute.value

        sheet.cell(row, column_index_from_string('B')).style = card_styles['blank_left']
        sheet.cell(row=row, column=column_index_from_string('C')).style = card_styles['txt_italic']
        sheet.cell(row=row, column=column_index_from_string('D')).style = card_styles['txt_ident']
        row += 1

    sheet.cell(row=row, column=column_index_from_string('B')).value = ''
    sheet.cell(row=row, column=column_index_from_string('C')).value = 'Параметры'
    sheet.cell(row=row, column=column_index_from_string('D')).value = ''

    sheet.cell(row, column_index_from_string('B')).style = card_styles['blank_left']
    sheet.cell(row=row, column=column_index_from_string('C')).style = card_styles['h2']
    sheet.cell(row=row, column=column_index_from_string('D')).style = card_styles['h2']
    sheet.cell(row, column_index_from_string('D')).border = card_borders['txt_c']

    for option in quote_data.options:
        row += 1
        sheet.cell(row=row, column=column_index_from_string('B')).value = ''
        sheet.cell(row=row, column=column_index_from_string('C')).value = option.name
        sheet.cell(row=row, column=column_index_from_string('D')).value = ''

        sheet.cell(row, column_index_from_string('B')).style = card_styles['blank_left']
        sheet.cell(row=row, column=column_index_from_string('C')).style = card_styles['h3']
        sheet.cell(row=row, column=column_index_from_string('D')).style = card_styles['h3']
        sheet.cell(row, column_index_from_string('D')).border = card_borders['txt_c']

        row += 1
        sheet.cell(row=row, column=column_index_from_string('B')).value = ''
        sheet.cell(row=row, column=column_index_from_string('C')).value = 'от'
        sheet.cell(row=row, column=column_index_from_string('D')).value = option.left

        sheet.cell(row, column_index_from_string('B')).style = card_styles['blank_left']
        sheet.cell(row=row, column=column_index_from_string('C')).style = card_styles['txt_italic']
        sheet.cell(row=row, column=column_index_from_string('D')).style = card_styles['txt_ident']

        row += 1
        sheet.cell(row=row, column=column_index_from_string('B')).value = ''
        sheet.cell(row=row, column=column_index_from_string('C')).value = 'до'
        sheet.cell(row=row, column=column_index_from_string('D')).value = option.right

        sheet.cell(row, column_index_from_string('B')).style = card_styles['blank_left']
        sheet.cell(row=row, column=column_index_from_string('C')).style = card_styles['txt_italic']
        sheet.cell(row=row, column=column_index_from_string('D')).style = card_styles['txt_ident']

        row += 1
        sheet.cell(row=row, column=column_index_from_string('B')).value = ''
        sheet.cell(row=row, column=column_index_from_string('C')).value = 'единица измерения'
        sheet.cell(row=row, column=column_index_from_string('D')).value = option.unit_measure

        sheet.cell(row, column_index_from_string('B')).style = card_styles['blank_left']
        sheet.cell(row=row, column=column_index_from_string('C')).style = card_styles['txt_italic']
        sheet.cell(row=row, column=column_index_from_string('D')).style = card_styles['txt_ident']

        row += 1
        sheet.cell(row=row, column=column_index_from_string('B')).value = ''
        sheet.cell(row=row, column=column_index_from_string('C')).value = 'шаг'
        sheet.cell(row=row, column=column_index_from_string('D')).value = option.step

        sheet.cell(row, column_index_from_string('B')).style = card_styles['blank_left']
        sheet.cell(row=row, column=column_index_from_string('C')).style = card_styles['txt_italic']
        sheet.cell(row=row, column=column_index_from_string('D')).style = card_styles['txt_ident']

        row += 1
        sheet.cell(row=row, column=column_index_from_string('B')).value = ''
        sheet.cell(row=row, column=column_index_from_string('C')).value = 'тип'
        sheet.cell(row=row, column=column_index_from_string('D')).value = option.option_type

        sheet.cell(row, column_index_from_string('B')).style = card_styles['blank_left']
        sheet.cell(row=row, column=column_index_from_string('C')).style = card_styles['txt_italic']
        sheet.cell(row=row, column=column_index_from_string('D')).style = card_styles['txt_ident']

    row += 1
    sheet.cell(row=row, column=column_index_from_string('B')).value = ''
    sheet.cell(row=row, column=column_index_from_string('C')).value = 'Связанные расценки'
    sheet.cell(row=row, column=column_index_from_string('D')).value = ''

    sheet.cell(row, column_index_from_string('B')).style = card_styles['blank_left']
    sheet.cell(row=row, column=column_index_from_string('C')).style = card_styles['h2']
    sheet.cell(row=row, column=column_index_from_string('D')).style = card_styles['h2']
    sheet.cell(row, column_index_from_string('D')).border = card_borders['txt_c']

    row += 1
    sheet.cell(row=row, column=column_index_from_string('B')).value = ''
    sheet.cell(row=row, column=column_index_from_string('C')).value = 'Тип связанной расценки'
    sheet.cell(row=row, column=column_index_from_string('D')).value = quote_data.type_quote

    sheet.cell(row, column_index_from_string('B')).style = card_styles['blank_left']
    sheet.cell(row=row, column=column_index_from_string('C')).style = card_styles['txt_italic']
    sheet.cell(row=row, column=column_index_from_string('D')).style = card_styles['txt_ident']

    row += 1
    sheet.cell(row=row, column=column_index_from_string('B')).value = ''
    sheet.cell(row=row, column=column_index_from_string('C')).value = 'Шифр родительской расценки'
    sheet.cell(row=row, column=column_index_from_string('D')).value = quote_data.parent_quote

    sheet.cell(row, column_index_from_string('B')).style = card_styles['blank_left']
    sheet.cell(row=row, column=column_index_from_string('C')).style = card_styles['txt_italic']
    sheet.cell(row=row, column=column_index_from_string('D')).style = card_styles['txt_ident']

    sheet.cell(row, column_index_from_string('B')).border = card_borders['last_b']
    sheet.cell(row, column_index_from_string('C')).border = card_borders['last_c']
    sheet.cell(row, column_index_from_string('D')).border = card_borders['last_d']
    return row+2


def write_excel_file(excel_file_name: str, quotes_data: list[Quote]):
    if file_unused(excel_file_name):
        book, sheet = None, None
        try:
            book = openpyxl.Workbook()
            DEFAULT_FONT.font = "Calibri"
            DEFAULT_FONT.sz = 11
            sheet = book.active  # получить ссылку на активную книгу
            sheet.title = "QuoteCards"
            # отключить сетку
            sheet.sheet_view.showGridLines = False
        except IOError as err:
            output_message_exit(f"Ошибка при создании excel файла: {excel_file_name!r}", f"{err}")

        _styles_init(book, card_styles)
        _basic_header_output(sheet)

        row = 2
        for quote in quotes_data:
            row = _quote_card_output(sheet, quote, row)

        if book:
            book.save(excel_file_name)
            book.close()
    else:
        output_message_exit(f"Файл занят другим приложением: ", f"{excel_file_name!r}")


if __name__ == "__main__":
    qd = [Quote(table='5.1-1-1-0-1', code='5.1-1-1',
                title='Генератор синхронный (компенсатор) напряжением до 1 кв, мощностью до 100 квт', measure='1 шт.',
                type_quote=None, parent_quote=None, attributes=[Attribute(name='Элемент', value='Генератор'),
                                                                Attribute(name='Тип', value='Синхронный/компенсатор')],
                options=[Option(name='Напряжение', left='0', right='1', unit_measure='кВ', step=None, option_type='3'),
                         Option(name='Мощность', left='0', right='100', unit_measure='кВт', step=None,
                                option_type='3')]), Quote(table='5.1-1-1-0-1', code='5.1-1-2',
                                                          title='Генератор синхронный (компенсатор) напряжением до 1 кв, мощностью свыше 100 квт',
                                                          measure='1 шт.', type_quote=None, parent_quote=None,
                                                          attributes=[Attribute(name='Элемент', value='Генератор'),
                                                                      Attribute(name='Тип',
                                                                                value='Синхронный/компенсатор')],
                                                          options=[Option(name='Напряжение', left='0', right='1',
                                                                          unit_measure='кВ', step=None,
                                                                          option_type='3'),
                                                                   Option(name='Мощность', left='100', right='-',
                                                                          unit_measure='кВт', step=None,
                                                                          option_type='3')])]



    write_excel_file(r"C:\Users\kazak.ke\Documents\PythonProjects\QuotesReportCard\output\test.xlsx", qd)
