# https://htmlcolorcodes.com/color-chart/flat-design-color-chart/

from openpyxl.styles import (Border, Side, Color, PatternFill, Font, Alignment, NamedStyle)
from icecream import ic

card_fonts = {
    'normal': Font(name='Calibri', color=Color(rgb='000000'), size=11, bold=False, italic=False),
    'normal_bold': Font(name='Calibri', color=Color(rgb='000000'), size=11, bold=True, italic=False),
    'normal_italic': Font(name='Calibri', color=Color(rgb='000000'), size=11, bold=False, italic=True),
    'bold_blue': Font(name='Calibri', color=Color(rgb='000000'), size=11, bold=True, italic=False), #'366092'
    'bold_italic_blue': Font(name='Calibri', color=Color(rgb='000000'), size=11, bold=True, italic=True), #'366092'
    'normal_gray': Font(name='Calibri', color=Color(rgb='00B3B6B7'), size=11, bold=False, italic=False),
}

card_borders = {
    "basic_left": Border(left=Side(border_style="thick", color=Color(rgb='A6A6A6')),
                        right=Side(border_style="thin", color=Color(rgb='A6A6A6'))
                         ),
    "basic_right": Border(left=None, right=Side(border_style="thick", color=Color(rgb='A6A6A6')), top=None,
                          bottom=None, ),
    "basic_top": Border(left=None, right=None, top=Side(border_style="thick", color=Color(rgb='A6A6A6')),
                        bottom=None, ),
    "basic_bottom": Border(left=None, right=None, top=None,
                           bottom=Side(border_style="thin", color=Color(rgb='A6A6A6'))),

    "txt_c": Border(left=None, right=Side(border_style="thick", color=Color(rgb='A6A6A6')), top=None,
                           bottom=Side(border_style="thin", color=Color(rgb='A6A6A6'))),

    "h1": Border(left=None, right=None,
                 top=Side(border_style="thick", color=Color(rgb='A6A6A6')),
                 bottom=Side(border_style="thin", color=Color(rgb='A6A6A6'))),

    "h1_c": Border(left=None, right=Side(border_style="thick", color=Color(rgb='A6A6A6')),
                 top=Side(border_style="thick", color=Color(rgb='A6A6A6')),
                 bottom=Side(border_style="thin", color=Color(rgb='A6A6A6'))),


    "table_code": Border(left=Side(border_style="thick", color=Color(rgb='A6A6A6')),
                         right=Side(border_style="thin", color=Color(rgb='A6A6A6')),
                         top=Side(border_style="thin", color=Color(rgb='A6A6A6')),
                         bottom=Side(border_style="thin", color=Color(rgb='A6A6A6'))),

    "quote_code": Border(left=Side(border_style="thick", color=Color(rgb='A6A6A6')),
                         right=Side(border_style="thin", color=Color(rgb='A6A6A6')),
                         top=Side(border_style="thick", color=Color(rgb='A6A6A6')),
                         bottom=Side(border_style="thin", color=Color(rgb='A6A6A6'))),

    "last_b": Border(left=Side(border_style="thick", color=Color(rgb='A6A6A6')),
                     right=Side(border_style="thin", color=Color(rgb='A6A6A6')),
                     top=None,
                     bottom=Side(border_style="thick", color=Color(rgb='A6A6A6'))),

    "last_c": Border(left=None, right=None, top=None,
                     bottom=Side(border_style="thick", color=Color(rgb='A6A6A6'))),

    "last_d": Border(left=None, right=Side(border_style="thick", color=Color(rgb='A6A6A6')),
                     top=Side(border_style="thin", color=Color(rgb='A6A6A6')),
                     bottom=Side(border_style="thick", color=Color(rgb='A6A6A6'))),

}

card_fills = {
    "h1": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00D0D3D4")),
    "h2": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00ECF0F1")),
    "h3": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00F7F9F9")),
    "blank": PatternFill(patternType=None, fill_type=None, fgColor=Color(rgb="00FFFFFF")),
}

# card_fills = {
#     "h1": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="007FB3D5")),
#     "h2": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00A9CCE3")),
#     "h3": PatternFill(patternType="solid", fill_type="solid", fgColor=Color(rgb="00D4E6F1")),
#     "blank": PatternFill(patternType=None, fill_type=None, fgColor=Color(rgb="00FFFFFF")),
# }



card_alignments = {
    "center": Alignment(horizontal='center', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "left_wrap": Alignment(horizontal='left', vertical='bottom', wrap_text=True, shrink_to_fit=False, indent=0),
    "left": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=0),
    "left_ident": Alignment(horizontal='left', vertical='bottom', wrap_text=False, shrink_to_fit=False, indent=1),

}

card_styles: dict[str: NamedStyle] = {
    "h1": NamedStyle(name="h1", font=card_fonts["normal"], border=card_borders["h1"],
                     fill=card_fills["h1"], alignment=card_alignments["center"]),

    "h2": NamedStyle(name="h2", font=card_fonts["bold_blue"], border=card_borders["basic_bottom"],
                     fill=card_fills["h2"], alignment=card_alignments["center"]),

    "h3": NamedStyle(name="h3", font=card_fonts["normal_bold"], border=card_borders["basic_bottom"],
                     fill=card_fills["h3"], alignment=card_alignments["left"]),

    "txt_ident": NamedStyle(name="txt_ident", font=card_fonts["normal"], border=card_borders["txt_c"],
                            fill=card_fills["blank"], alignment=card_alignments["left_ident"]),

    "txt_italic": NamedStyle(name="txt_italic", font=card_fonts["normal_italic"], border=card_borders["basic_bottom"],
                             fill=card_fills["blank"], alignment=card_alignments["left"]),

    "txt_gray": NamedStyle(name="txt_gray", font=card_fonts["normal_gray"], border=card_borders["table_code"],
                           fill=card_fills["blank"], alignment=card_alignments["center"]),

    "blank_left": NamedStyle(name="blank_left", font=card_fonts["normal"], border=card_borders["basic_left"],
                             fill=card_fills["blank"], alignment=card_alignments["center"]),
}

if __name__ == '__main__':
    for style in card_styles:
        ic(style, type(style))
        ic(card_styles[style].name)
